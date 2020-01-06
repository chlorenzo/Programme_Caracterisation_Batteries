# Ce programme permet de réaliser une série de cycles de charge et décharge sur un accumulateur de manière autonome.
# Le document Article_3EI.docx présente en détail le fonctionnement du montage complet. 
# Ce code python est à lancer à partir du CPU, lui-même connecté au microcontroleur F28377S.

#Ce script détermine et transmet la consigne de la source de courant du montage. Pour cela, il récupère les données de tension et de courant sur le port série, et à partir de ces données détermine la valeur de la consigne de la source de courant. Ces données récupérées et calculées sont stockées dans un fichier excel.
#Ce ficher excel est appelé DonneesBatteries.xlsx et se situe dans le dossier PythonTests à la racine.
#Le chemin est à modifier si nécessaire.
#Quatre librairies sont à télecharger : pyserial, binascii, openpyxl et time.
#La première permet la lecture sur le port série, la seconde permet de convertir facilement la mesure en entier, openpyxl permet de stocker les données dans un fichier excel et la time permet d'inserer la variable de temps dans le relevé des données (nécessaire pour déterminer l'état de charge notamment).

#Les données envoyées sont transmises aux sorties analogiques DACA et DACB. Ces deux sorties créent la différence de potentiel DACA-DACB qui sert de consigne à la source de courant. 
#Dans le cas d'une décharge, on imposera une différence de potentiel positive soit DACA > DACB. Dans le cas de la décharge, la différence de potentiel doit être négative soit DACB > DACA.
#On imposera dans chaque situation une sortie nulle et l'autre positive. 


#Bibliothèques à importer.
#pyserial, binascii, openpyxl et time sont à installer.
# Os présent par défaut.
import serial
import os
import binascii
import time
from openpyxl import Workbook
os.chdir("C:/PythonTests")      #choix du répertoir de travail.

#Détermination paramètres de la simulation. Ces paramètres dépendent notamment de la technologie de la batterie, et du profil de la sollicitation à imposer.
V_max = 4200	#Tension max de l'accumulateur (en mV)
V_min = 3000    #Tension min de l'accumulateur (en mV)
I_min_Charge = 10 #Courant limite à partir duquel on stoppe la charge (en mA)
I_max_Decharge =  -10 #Courant limite à partir duquel on stoppe la décharge (en mA)
TempsSimu = 3*3600 	#Durée de la simulation (en s)
I_Charge = 650 #Courant de charge, (en mA)   
I_Decharge = 650 #Courant de décharge, (en mA) (valeur absolue)
I_Incr_Dech = 20 #Incrément de réduction du courant, en mA, pour maintenir la tension constante, en décharge (en mA).
I_Incr_Ch = 20 #Incrément de réduction du courant, en mA, pour maintenir la tension constante, en charge (en mA) (valeur absolue).
R_shunt = 0.1 #Valeur de la resistance de shunt (en Ohm) sur le montage hardware.


#Le conversion réalisée par le DAC impose la pour relation 1 bit = 40 mV de consigne.
#La source de courant fournit un courant image de la tension de consigne : I = V/10.
#Le système impose une relation 1 bit envoyé = 4 mA fournit par la source. 
I_Charge_bit = int(I_Charge/4)
I_Decharge_bit = int(I_Decharge/4)
I_Incr_Dech_bit = int(I_Incr_Dech/4)
I_Incr_Ch_bit = int(I_Incr_Ch/4)


#Temps initial pris au début de l'expérience.
T_init = time.time()

#Stockage des données sur la feuille excel : définition des titres dans les cases correspondantes.
wb = Workbook()
ws = wb.active
ws['A1'] = "Données de tension et de tension shunt " 
ws['A2'] = "Time (s)"
ws['B2'] = "BusVoltage (V)"
ws['C2'] = "ShuntVoltage (V)"
ws['D2'] = "SourceCommande A"
ws['E2'] = "SourceCommande A"

i = 0
j = 0

#Definition du port serie à lire ainsi que de la vitesse de transmisson.
#115200 bauds pour synchronisme avec launchpad.
ser = serial.Serial('COM12', baudrate = 115200)                                    


#On commence par charger la batterie à 1C. CourantCharge = 650 mAh
#Initialisation : ici on commence par décharger l'accumulateur. Si l'on souhaite commencer par une charge, alors il faut imposer 0 au DAC_A et une valeur non nulle au DAC_B. 
DAC_A = bytes([I_Decharge_bit])            #On convertit la valeur souhaitée en byte, pour l'envoyer avec le bon format.
DAC_B = bytes([0])              #On convertit la valeur souhaitée en byte. 

#L'expérience dure tant que le temps de cyclage imposé n'est pas atteint.
while time.time() -  T_init < TempsSimu :                                                                    

    i+=1
    donneesRecues = ser.read(4)                       #Lecture des 4 bytes reçus sur le port série
    TensionBusByte = donneesRecues[:2]				 #Les deux premiers bytes sont attribués à la variable TensionBus
    TensionShuntByte = donneesRecues[2:4]             #Les deux bytes suivants sont correspondent à la tension de la resistance de shunt  

	#Conversion des données dans le bon format pour pouvoir les stocker.
    TensionBusHexa = binascii.hexlify(TensionBusByte) 
    TensionShuntHexa = binascii.hexlify(TensionShuntByte)
    TensionBusInt = int(TensionBusHexa, 16)       
    TensionShuntInt = int(TensionShuntHexa, 16)      

	#La tensionShunt est négative dans le cas de la charge de l'accumulateur. 
	#La donnée récupérée est codée en complément à 2. Il faut donc la convertir.
	#Si le bit de poids fort de la tension shunt vaut 1, cela signifie que la tension est négative. 
	#Pour convertir, on retranche 2^16 à la tension shunt et on obtient la tension négative correspondante.
    if TensionShuntHexa[15] == 1 :
        TensionShuntInt = (TensionShuntInt - 2**16)

	#La tension du bus est déterminée avant la resistance de shunt. 
	#A la valeur mesurée, il faut donc retrancher la tension de cette resistance pour déterminer la tension du bus.
	#La tension shunt est indiquée en 10 micro-ohm. Il faut donc la multiplier par 0.01 pour obtenir des mV.
    TensionBusInt -= TensionShuntInt*0.01
	I_Bus_Int = TensionShuntInt*0.01/R_shunt
    
	#Conversion des données en chaines de charactère pour pouvoir le stockage dans le fichier excel.
    TensionBus = str(TensionBusInt)
    I_Bus = str(I_Bus_Int)            
    
	#Récupération de la donnée de temps
    TempsInt = time.time()
    Temps = round(TempsInt - T_init, 2)
  
	#On surveille tout d'abord la tension de l'accumulateur.
	#Dès que la tension mesuree depasse la tension Seuil max :  on réduit le courant, par incrément du nombre de bit qui correspondent à la valeur de I_Incr_Dech (quelques mA).

	#######################################################
	### 												###
	### 	Algorithme de cyclage de l'accumulateur 	###
	###													###
	#######################################################
	
	#Il a été choisi de charger à courant constant jusqu'à ce que l'accumulateur atteigne une tension Vmax. Une fois cette tension atteinte, la charge est maintenue à tension constante.
	#Le courant est réduit par incrément définit par l'utilisateur. Une fois un seuil en courant atteint, la phase de décharge est déclenchée. 
	#La décharge est définie de la même manière (courant constant, puis tension constante avec réduction du courant par incréments constants).
	#Cette méthode est généralement utilisée pour recharger les batteries lithium-ion.
	#En modifiant la partie ci-dessous, l'utilisateur peut définir la manière de cycler l'accumulateur.
	
	
	#Lors des phases de charge ou de décharge à courant constant, la valeur des DAC est maintenue et ne varie pas.
	
	#En charge, dès que la tension mesurée dépasse la tension max, la consigne sur le DAC A est réduite.
    if TensionBusInt >= V_max :  
		#Tant que la tension le courant n'a pas atteint son seuil min ou que sa décroissance ne va pas entrainer un passage à la décharge : on réduit la consigne de la source de courant.
        if I_Bus_Int >= I_max_Charge and int(binascii.hexlify(DAC_A),16) > I_Incr_Ch_bit:
            DAC_AHexa = binascii.hexlify(DAC_A)
            DAC_AInt = int(DAC_AHexa, 16)
            DAC_A = bytes([DAC_AInt - I_Incr_Ch_bit])
            DAC_B = bytes([0])
        #Si une des deux conditions n'est pas respectée on passe à la decharge.
        else :
            DAC_A = bytes([0])
            DAC_B = bytes([I_Decharge_bit])
	#En decharge 
		#Si on dépasse le seuil min de la tension : on augmente le courant.

	#En décharge, dès que la tension mesurée est inférieure à la tension min, la consigne envoyée sur le DAC B est réduite. 
    elif TensionBusInt <= V_min:
		#Tant que la tension le courant n'a pas atteint son seuil min ou que sa décroissance ne va pas entrainer un passage à la charge : on réduit la consigne de la source de courant (en valeur absolue).
        if I_Bus_Int <= I_max_Decharge and int(binascii.hexlify(DAC_B),16) > I_Incr_Dech_bit :
            DAC_BHexa = binascii.hexlify(DAC_B)
            DAC_BInt = int(DAC_BHexa, 16)
            DAC_B = bytes([DAC_BInt - I_Incr_Dech_bit])
            DAC_A = bytes([0])
    #Si une des deux conditions n'est pas respectée on passe à la decharge.        
	else :
            DAC_A = bytes([I_Charge_bit])
            DAC_B = bytes([0])

	#On envoit le sur le port série les bytes de commande des deux DAC.
    ser.write(DAC_A)                           
    ser.write(DAC_B)
	
	#Conversion des données en chaîne de caractère
    DAC_A_STR = str(int(binascii.hexlify(DAC_A),16))
    DAC_B_STR = str(int(binascii.hexlify(DAC_B),16))
	
	#On stock dans le fichier les données, ligne par ligne. 
	#A partir de ces données, on peut déterminer notamment l'évolution du de l'état de charge de l'accumulateur. 
    ws.append([Temps, TensionBus, I_Bus, DAC_A_STR, DAC_B_STR])                           

wb.save("DonneesBatteries.xlsx") #A la fin de la boucle, le fichier est sauvegardé.
ser.close()                      # On ferme ensuite le fichier.


